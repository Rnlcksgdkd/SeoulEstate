"""
엑셀 파일 처리 모듈
pandas와 openpyxl을 사용하여 견적서를 엑셀 파일로 변환합니다.
"""

import os
import logging
from typing import Dict, Any
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.config.constants import TEMPLATE_DIR, DEFAULT_TEMPLATE, ERROR_MESSAGES

logger = logging.getLogger(__name__)

class ExcelHandler:
    """엑셀 파일 처리 클래스"""

    def __init__(self):
        """템플릿 디렉토리 확인 및 생성"""
        try:
            if not os.path.exists(TEMPLATE_DIR):
                os.makedirs(TEMPLATE_DIR)
        except Exception as e:
            logger.error(f"템플릿 디렉토리 생성 중 오류 발생: {e}")
            raise Exception(ERROR_MESSAGES["FILE_ERROR"])

    def export_estimate(self, estimate_data: Dict[str, Any], output_path: str) -> bool:
        """
        견적 데이터를 엑셀 파일로 내보냅니다.
        
        Args:
            estimate_data: 견적 데이터
            output_path: 출력할 엑셀 파일 경로
            
        Returns:
            bool: 성공 여부
        """
        try:
            # 데이터프레임 생성
            df_labor = self._create_labor_cost_df(estimate_data.get('labor_costs', []))
            df_other = self._create_other_costs_df(estimate_data)
            
            # 엑셀 파일 생성
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 인건비 시트
                df_labor.to_excel(writer, sheet_name='인건비', index=False)
                self._apply_styles(writer.sheets['인건비'])
                
                # 기타 비용 시트
                df_other.to_excel(writer, sheet_name='기타비용', index=False)
                self._apply_styles(writer.sheets['기타비용'])
                
                # 열 너비 자동 조정
                for sheet in writer.sheets.values():
                    for column in sheet.columns:
                        max_length = 0
                        column = column[0].column_letter
                        for cell in sheet[column]:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column].width = adjusted_width
            
            return True
            
        except Exception as e:
            logger.error(f"엑셀 파일 생성 중 오류 발생: {e}")
            raise Exception(ERROR_MESSAGES["FILE_ERROR"])

    def _create_labor_cost_df(self, labor_costs: list) -> pd.DataFrame:
        """인건비 데이터프레임을 생성합니다."""
        data = []
        for cost in labor_costs:
            data.append({
                '역할': cost['role'],
                '단가(월)': cost['monthly_rate'],
                '투입기간(월)': cost['duration'],
                '금액': cost['monthly_rate'] * cost['duration']
            })
        return pd.DataFrame(data)

    def _create_other_costs_df(self, estimate_data: Dict[str, Any]) -> pd.DataFrame:
        """기타 비용 데이터프레임을 생성합니다."""
        data = [
            {'항목': '개발 환경 구축 비용', '금액': estimate_data.get('setup_cost', 0), '비고': ''},
            {'항목': '라이선스 및 외부 서비스', '금액': estimate_data.get('license_cost', 0), '비고': ''},
            {'항목': '유지보수 비용', '금액': estimate_data.get('maintenance_cost', 0), '비고': '옵션'},
            {'항목': '예비비', '금액': estimate_data.get('contingency', 0), '비고': '전체 금액의 10%'},
            {'항목': '총 견적 금액', '금액': estimate_data.get('total_cost', 0), '비고': ''}
        ]
        return pd.DataFrame(data)

    def _apply_styles(self, worksheet) -> None:
        """워크시트에 스타일을 적용합니다."""
        # 스타일 정의
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center')
        
        # 헤더 스타일 적용
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = alignment
            
        # 데이터 셀 스타일 적용
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = alignment

    def load_template(self) -> pd.DataFrame:
        """기본 템플릿을 로드합니다."""
        template_path = os.path.join(TEMPLATE_DIR, DEFAULT_TEMPLATE)
        try:
            return pd.read_excel(template_path)
        except Exception as e:
            logger.error(f"템플릿 로드 중 오류 발생: {e}")
            raise Exception(ERROR_MESSAGES["FILE_ERROR"]) 